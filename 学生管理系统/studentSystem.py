import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import *
from tkinter.ttk import *
import inspect
import openpyxl


# 登陆界面
def form1():
    # 生成tk窗口，设置居中显示，并且设置无法拉伸
    root = tk.Tk()
    root.title("学生管理系统登录页面")
    width = 300
    height = 200
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int(screen_width / 2 - width / 2)
    y = int(screen_height / 2 - height / 2)
    size = '{}x{}+{}+{}'.format(width, height, x, y)
    root.geometry(size)
    root.resizable(False, False)

    # 进行登录页面编辑
    tk.Label(root, text="用户名:", font=('微软雅黑', 20, 'bold'),
             anchor='w', pady=15, padx=10).grid(row=0, sticky='w')
    tk.Label(root, text="密  码:", font=('微软雅黑', 20, 'bold'),
             anchor='w', pady=15, padx=10).grid(row=1, sticky='w')
    # 设置账号输入框
    user_entry = tk.Entry(root)
    user_entry.grid(row=0, column=1)
    # 设置密码输入框
    psw_entry = tk.Entry(root, show='*')
    psw_entry.grid(row=1, column=1)

    def login():
        username = user_entry.get()
        password = psw_entry.get()
        if username == "admin" and password == "admin":
            messagebox.showinfo(title='登录成功', message="successful")
            root.destroy()
            form2()

        else:
            messagebox.showinfo(title='登录失败', message="fail")
            psw_entry.delete(0, 'end')
            user_entry.delete(0, "end")

    def logout():
        result = messagebox.askyesno("提示", message="你确定要关闭窗口吗")
        if result:
            root.quit()

    tk.Button(root, text='登录', width=10, command=login).grid(row=3, column=0, columnspan=1, sticky='w', padx=30,
                                                               pady=5)
    tk.Button(root, text='退出', width=10, command=logout).grid(row=3, column=1, columnspan=1, sticky='e', padx=10,
                                                                pady=5)

    root.mainloop()


def form2():
    root = tk.Tk()
    root.title("学生管理系统管理页面")
    width = 460
    height = 380
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int(screen_width / 2 - width / 2)
    y = int(screen_height / 2 - height / 2)
    size = '{}x{}+{}+{}'.format(width, height, x, y)
    root.geometry(size)
    root.resizable(False, False)

    def show():
        pass

    def onclick():  # 打开文件
        path = inspect.getfile(inspect.currentframe()).split("\\")
        path = "\\".join(path[:-1])
        file_name = filedialog.askopenfilename(title="打开我的文件", initialdir=path,
                                               filetypes=[("XLSX工作表", ".xlsx"), ('All Files', '*')])
        # file_name是文件的路径名称
        workbook = openpyxl.load_workbook(file_name)
        # 获取第一个sheet表格
        table = workbook.active
        # 获取每一行数据
        rows = list(table.rows)[1:]
        i = 1
        # 显示每一行数据
        for row in rows:
            row_list = []
            for c in row:
                row_list.append(c.value)
            TreeViewtree.insert("", i, values=row_list)
            i += 1
        workbook.close()

    def delete():
        iid = TreeViewtree.selection()
        name = TreeViewtree.item(iid[0], "values")[1]

        path = inspect.getfile(inspect.currentframe()).split("\\")
        path = "\\".join(path[:-1]) + '\\学生信息.xlsx'
        # file_name是文件的路径名称
        workbook = openpyxl.load_workbook(path)
        # 获取第一个sheet表格
        table = workbook.active
        max_row_num = table.max_row

        row_n = 2
        col_n = 2
        while row_n <= max_row_num:
            if table.cell(row=row_n, column=col_n).value == name:
                table.delete_rows(row_n)
                break
            else:
                row_n += 1
        TreeViewtree.delete(iid[0])
        workbook.save(filename='学生信息.xlsx')
        workbook.close()

    def add():
        add_win = tk.Tk()
        add_win.title("学院添加")
        add_win.geometry('710x100+500+400')
        add_win.resizable(False, False)
        # 学号
        tk.Label(add_win, text='学号', font="微软雅黑 13").grid(row=1, column=1)
        id_ent = tk.Entry(add_win, width=10)
        id_ent.grid(row=1, column=2)
        # 姓名
        tk.Label(add_win, text='姓名', font="微软雅黑 13").grid(row=1, column=3)
        name_ent = tk.Entry(add_win, width=10)
        name_ent.grid(row=1, column=4)
        # 性别
        tk.Label(add_win, text='性别', font="微软雅黑 13").grid(row=1, column=5)
        sex_ent = tk.Entry(add_win, width=10)
        sex_ent.grid(row=1, column=6)
        # 年龄
        tk.Label(add_win, text='年龄', font="微软雅黑 13").grid(row=1, column=7)
        age_ent = tk.Entry(add_win, width=15)
        age_ent.grid(row=1, column=8)
        # 手机号
        tk.Label(add_win, text='手机号', font="微软雅黑 13").grid(row=1, column=9)
        iphone_ent = tk.Entry(add_win, width=15)
        iphone_ent.grid(row=1, column=10)

        def add_student():
            path = inspect.getfile(inspect.currentframe()).split("\\")
            path = "\\".join(path[:-1]) + '\\学生信息.xlsx'
            # file_name是文件的路径名称
            workbook = openpyxl.load_workbook(path)
            # 获取第一个sheet表格
            table = workbook.active
            iid = id_ent.get()
            name = name_ent.get()
            sex = sex_ent.get()
            age = age_ent.get()
            iphone = iphone_ent.get()
            info_list = [iid, name, sex, age, iphone]
            TreeViewtree.insert("", END, values=info_list)
            row_num = table.max_row
            for i in range(len(info_list)):
                table.cell(row=row_num, column=i + 1).value = info_list[i]
            print(info_list)
            workbook.save(filename='学生信息.xlsx')
            workbook.close()

        tk.Button(add_win, text="添加", command=add_student, font="微软雅黑 13").grid(row=2, column=1, padx=10, pady=20,
                                                                                      sticky='e')
        tk.Button(add_win, text="退出", command=add_win.destroy, font="微软雅黑 13").grid(row=2, column=2, padx=10,
                                                                                          pady=20, sticky='w')
        add_win.mainloop()

    def update():
        pass

    def find():
        pass

    # Treeview控件
    frame01 = Frame(root)
    frame01.place(x=10, y=10, width=420, height=220)
    # 加载滚动条
    scrollBar = Scrollbar(frame01)
    scrollBar.pack(side=RIGHT, fill=Y)
    # 准备表格
    TreeViewtree = Treeview(frame01, columns=("学号", "姓名", "性别", "年龄", "手机号"), show="headings",
                            yscrollcommand=scrollBar.set)
    # 设置每一列的宽度和对齐方式
    TreeViewtree.column("学号", width=80, anchor="center")
    TreeViewtree.column("姓名", width=80, anchor="center")
    TreeViewtree.column("性别", width=60, anchor="center")
    TreeViewtree.column("年龄", width=60, anchor="center")
    TreeViewtree.column("手机号", width=120, anchor="center")
    # 设置表头的标题文本
    TreeViewtree.heading("学号", text="学号")
    TreeViewtree.heading("姓名", text="姓名")
    TreeViewtree.heading("性别", text="性别")
    TreeViewtree.heading("年龄", text="年龄")
    TreeViewtree.heading("手机号", text="手机号")
    # 设置关联
    scrollBar.config(command=TreeViewtree.yview)
    # 加载表格信息
    TreeViewtree.pack()
    # 插入数据
    # for i in range(10):  # i 是索引
    #     TreeViewtree.insert("", i, values=["9500" + str(i), "张三", "男", "23", "15622338793"])  # 展示

    frame02 = Frame(root)
    frame02.place(x=10, y=240)
    tk.Label(frame02, text='选择学生名单(Excel)', font="微软雅黑 13").grid(row=1, column=0, sticky='e')
    tk.Button(frame02, text="浏览", command=onclick, font="微软雅黑 13").grid(row=1, column=1, sticky='ns')
    showname = tk.Label(frame02, text=' ', font="微软雅黑 13", fg='blue')
    showname.grid(row=8, column=0, columnspan=2, padx=30, sticky='ns')

    # 增删改查按钮选项
    frame03 = Frame(root)
    frame03.place(x=10, y=300)
    tk.Button(frame03, text="添加", command=add, font="微软雅黑 13").grid(row=1, column=3, padx=20, sticky='e')
    tk.Button(frame03, text="删除", command=delete, font="微软雅黑 13").grid(row=1, column=5, padx=20, sticky='e')
    tk.Button(frame03, text="修改", command=update, font="微软雅黑 13").grid(row=1, column=7, padx=20, sticky='e')
    tk.Button(frame03, text="退出", command=root.destroy, font="微软雅黑 13").grid(row=1, column=9, padx=20, sticky='e')

    root.mainloop()


if __name__ == '__main__':
    form1()
